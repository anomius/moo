"""OCCP Business Constraints Tool - Streamlit Application.

This module provides a Streamlit-based UI for configuring and submitting business
constraints for OCCP cycles, including integration with Snowflake,
email notifications, and Excel report generation.
"""

import os
import sys
import re
import warnings
import datetime
from pathlib import Path
from itertools import combinations, count
from typing import List
from io import BytesIO
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import base64
import json

# Third-party imports
import streamlit as st
from streamlit import session_state as ss
import pandas as pd
import requests
import yaml
from dateutil import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from snowflake.connector.pandas_tools import write_pandas

# Local application/library specific imports
from turing_generic_lib.utils.config import TuringConfig

ENV = os.environ.get("ENVIRONMENT", "DEV")
from utils.utilization_automation import channel_utilization_status
from utils.utils import (
    get_brand_code,
    SnowflakeConnection,
)
from utils.output_mapping import create_business_constraints_file

from collections import defaultdict
from itertools import chain

# === CONSTANTS FOR DUPLICATED LITERALS ===
HISTORICAL_INTERACTION_LEVEL = "Historical Interaction level"
HCP_SEGMENT_LEVEL = "HCP Segment level"
OCCP_CONTEXT = "OCCP Context"
SALES_REP_CONSTRAINTS = "Sales Rep. constraints"
MULTIBRAND_OCCP = "For Multibrand OCCP only:"
TYPE_LABEL = "HCP Constraints"
CHOOSE_O = "Choose an option"
date_format_str = "%Y-%m-%d"
warnings.filterwarnings("ignore")

SCRIPT_DIR = os.path.dirname(__file__)
CONFIG_DIR = os.path.join(SCRIPT_DIR, "config")
SQL_DIR = os.path.join(SCRIPT_DIR, "sql")
sys.path.append(os.path.join(SCRIPT_DIR, "../"))

style_path = os.path.join(SCRIPT_DIR, "utils", "style.txt")
config_path = Path(__file__).parent / "config" / "snowflake_config.yaml"
email_path = os.path.join(CONFIG_DIR, "email_config.yaml")
brand_mapping_path = os.path.join(CONFIG_DIR, "brand_mapping.yaml")
with open(brand_mapping_path, "r", encoding="utf-8") as file:
    brand_map = yaml.safe_load(file)["BRAND"]

class OCCPTool:
    """Main class for OCCP Business Constraints Tool."""

    def __init__(self):
        """Initialize OCCPTool with configuration and Snowflake connection."""
        self.config = TuringConfig(
            config_dir=Path(__file__).parents[1] / "config",
            gbu="gen",
            countrycode="IT",
            brand="TOJ",
        )
        self.config.load()
        self.snowflake_conn, self.output_snowflake_con = SnowflakeConnection(
            self.config
        ).get_connection()

    def open_sql_file(self, sql_path: str) -> str:
        """open a .sql file and get the query as a string

        :param sql_path: {str} -- path to the .sql file
        :return: sql query as string
        """
        with open(sql_path, "r", encoding="utf-8") as file_read:
            query = file_read.read()
        return query

    def fetch_team_data_from_snowflake(self):
        """Fetch team and brand data from Snowflake."""
        country_sales_line_query = self.open_sql_file(
            f"{SQL_DIR}/country_sales_line.sql"
        )
        gbu_brand_query = self.open_sql_file(f"{SQL_DIR}/gbu_brand.sql")
        try:
            df_Team = pd.read_sql(country_sales_line_query, self.snowflake_conn)
            df_Team.dropna(subset=["COUNTRY", "SALES_TEAM"], inplace=True)
            df_Brand = pd.read_sql(gbu_brand_query, self.snowflake_conn)
            df_Brand.dropna(subset=["GLOBAL_BRAND"], inplace=True)
        finally:
            pass
        return df_Team, df_Brand

    def fetch_channel_data_from_snowflake(self):
        """Fetch channel data from Snowflake."""
        channel_query = self.open_sql_file(f"{SQL_DIR}/channel.sql")
        try:
            df_channel = pd.read_sql(channel_query, self.snowflake_conn)
            df_channel.dropna(subset=["CHANNEL"], inplace=True)
        finally:
            pass
        return df_channel

    def fetch_master_and_time_dim_data_from_snowflake(self):
        """Fetch master and time dimension data from Snowflake."""
        master_query = self.open_sql_file(f"{SQL_DIR}/master_data.sql")
        time_dim_query = self.open_sql_file(f"{SQL_DIR}/time_dimension.sql")
        try:
            df_master = pd.read_sql(master_query, self.snowflake_conn)
            df_time_dim = pd.read_sql(time_dim_query, self.snowflake_conn)
        finally:
            pass
        return df_master, df_time_dim

    def fetch_validate_data(self):
        try:
            e_consent_query = self.open_sql_file(
                f"{SQL_DIR}/e_consent_characterstics.sql"
            )
            rep_occp_query = self.open_sql_file(f"{SQL_DIR}/rep_occp.sql")
            characterstics_query = self.open_sql_file(
                f"{SQL_DIR}/characterstics_brand.sql"
            )
            self.convert_brand_names()
            e_consent_query = self.replace_query_params(query=e_consent_query)
            rep_occp_query = self.replace_query_params(query=rep_occp_query)
            e_consent_df = pd.read_sql(e_consent_query, self.snowflake_conn)
            rep_occp_df = pd.read_sql(rep_occp_query, self.snowflake_conn)
            characterstics_df = self.prepare_hcp_char(characterstics_query)
            return e_consent_df, rep_occp_df, characterstics_df
        except Exception as ex:
            st.error(f"Error fetching validation data: {ex}")
            return None, None, None

    def convert_brand_names(self):
        result = []
        for item in ss["brands"]:
            for _, row in ss.df_Brand.iterrows():
                full_name = (
                    f"{row['GLOBAL_BRAND']} {row['INDICATION_NAME']}".strip().upper()
                )
                if item.strip().upper() == full_name:
                    result.append(f"{row['GLOBAL_BRAND']} {row['INDICATION_CD']}")
                    break
            else:
                result.append(item)
        ss["brand_map"] = result
        return result

    def replace_query_params(self, query):
        cycle_end_date = (ss.cycle_start_date - datetime.timedelta(days=1)).strftime(
            date_format_str
        )
        query = query.replace("_COUNTRY_", f"'{ss['country']}'")
        query = query.replace("_CYCLE_END_DT_", f"'{cycle_end_date}'")
        if "_BRAND_" in query and "BRAND_NUM_" not in query:
            brand_str = ", ".join([f"'{b.strip().upper()}'" for b in ss["brand_map"]])
            query = query.replace("_BRAND_", f"{brand_str}")
        elif "BRAND_NUM_" in query:
            brand_num = ss["brand_id"] + 1
            query = query.replace("BRAND_NUM_", f"BRAND{brand_num}")
            query = query.replace(
                "_BRAND_", f"'{ss['brand_map'][ss['brand_id']].strip().upper()}'"
            )
        return query

    def prepare_hcp_char(self, query):
        hcp_char_df = None
        for brand_id, brand in enumerate(ss["brand_map"]):
            ss["brand_id"] = brand_id
            ss["brand"] = brand
            hcp_characteristic_map_final_query = self.replace_query_params(query)
            hcp_characteristic_map_df_brand = pd.read_sql(
                hcp_characteristic_map_final_query, self.snowflake_conn
            )
            if hcp_char_df is None:
                hcp_char_df = hcp_characteristic_map_df_brand
            else:
                hcp_char_df = pd.merge(
                    hcp_char_df,
                    hcp_characteristic_map_df_brand,
                    on=["HCP_ID"],
                    how="left",
                )
        return hcp_char_df

    def initialize_session_state(self):
        """Initialize session state variables."""
        session_vars = [
            "countrycode",
            "country",
            "brands",
            "cycle_name",
            "sales_team",
            "monobrand_channel",
            "multibrand_channel",
            "features",
            "constraints",
            "rep_capacity",
            "selected_envelope_toggle",
            "envelopeCriteria",
            "econsent_rte",
            "veeva_checkbox",
            "is_non_prescriber",
            "non_prescribers_priority",
        ]
        for var in session_vars:
            if var not in ss:
                ss[var] = "" if var in ["countrycode", "country", "cycle_name"] else {}

    def set_page_styling(self):
        """Set custom CSS and footer for the Streamlit page."""
        with open(style_path, "r", encoding="utf-8") as style_file:
            style_text = style_file.read()
        st.markdown(
            style_text,
            unsafe_allow_html=True,
        )

    def configure_page(self, obj):
        """Configure Streamlit page layout and load initial data."""
        st.set_page_config(
            page_title="OCCP Tool",
            layout="wide",
            page_icon=":shark:",
            initial_sidebar_state="expanded",
        )

        container = st.empty()
        container.markdown("\n\n\n")

        st.markdown(
            "<style>div.block-container{padding-top:1rem;}</style>",
            unsafe_allow_html=True,
        )

        self.set_page_styling()

        image_path = os.path.join(SCRIPT_DIR, "./utils/turing_logo.PNG")

        # Function to load and encode the image in base64
        def get_base64_image(image_path):
            """Load an image and return it as a base64-encoded string."""
            with open(image_path, "rb") as img_file:
                return base64.b64encode(img_file.read()).decode()

        # Get the base64-encoded image
        base64_image = get_base64_image(image_path)

        # Top Header Bar with embedded base64 image
        st.markdown(
            (
                "<div class='top-header'>"
                "<h1>OCCP Business Constraints Tool</h1>"
                f"<img src='data:image/png;base64,{base64_image}' alt='Turing Logo'>"
                "</div>"
                "<div class='main-content'>"
            ),
            unsafe_allow_html=True,
        )
        ss.df_Team, ss.df_Brand = obj.fetch_team_data_from_snowflake()

    def get_master_country_list(self, df_team) -> list[str]:
        """Return alphabetically‐sorted list of unique countries plus any extras."""
        countries = sorted(df_team["COUNTRY"].dropna().unique())
        if ss.country not in countries:
            ss.country = countries[0]
        return sorted(countries)

    def select_country(self, countries: list[str]) -> str:
        """Render the country selector and return the user choice."""
        return st.selectbox(
            "Select the Country Name",
            countries,
            index=countries.index(st.session_state.country),
            help=(
                "Select the country for which the Omni Channel Call Plan (OCCP) "
                "needs to be generated for the upcoming cycle."
            ),
        )

    def update_sales_team_options(self, df_team, country: str) -> None:
        """Refresh sales‐team options when country changes and handle selection."""
        if "options" not in st.session_state or ss.country != st.session_state.get(
            "last_country"
        ):
            self._set_sales_team_options(df_team, country)
        self._handle_sales_team_selection()

    def _set_sales_team_options(self, df_team, country: str):
        """Set sales team options for the selected country."""
        sales_teams = ss.df_Team[ss.df_Team["COUNTRY"] == ss.country][
            "SALES_TEAM"
        ].unique()
        sorted_sales_teams = sorted(
            [team for team in sales_teams if isinstance(team, str)],
            key=lambda x: x.lower(),
        )
        sorted_sales_teams.insert(0, CHOOSE_O)
        st.session_state.options = sorted_sales_teams
        st.session_state.last_country = ss.country

    def _handle_sales_team_selection(self):
        """Handle the sales team selection UI and validation."""
        options = st.session_state.options + ["Others"]
        selected_team = st.selectbox(
            "Select the Sales Line Name",
            options,
            label_visibility="visible",
            help="Select the sales line name for the upcoming Omni Channel Call Plan (OCCP) cycle for the chosen country.",
        )
        if selected_team == CHOOSE_O:
            st.write("Please select a valid option.")

        ss.team_flag = False
        if selected_team == "Others":
            new_team = st.sidebar.text_input(
                "If 'Others', please specify the Sales Line Name."
            )
            if new_team:
                if re.match(r"^[A-Z]{2}_.+$", new_team):
                    if new_team not in st.session_state.options:
                        ss.sales_team = new_team
                        ss.team_flag = True
                    else:
                        st.warning("This Sales line already exists.")
                else:
                    st.error(
                        "Sales line must follow the pattern: COUNTRYCODE_<Sales_Line_Name>"
                    )
            else:
                st.error("Please enter a valid Sales line.")
        else:
            ss.team_flag = True
            ss.sales_team = selected_team

    @staticmethod
    def sales_team_selector(options: list[str]) -> str:
        """Render sales-team select box and return the choice."""
        return st.selectbox(
            "Select the Sales Line Name",
            options + ["Others"],
            help=(
                "Select the sales line name for the upcoming Omni Channel Call "
                "Plan (OCCP) cycle for the chosen country."
            ),
        )

    @staticmethod
    def handle_new_sales_team(choice: str) -> tuple[bool, str]:
        """Validate / capture new sales team if user picked 'Others'."""
        if choice != "Others":
            return True, choice  # valid existing option

        # Ask user for free-text input
        new_team = st.sidebar.text_input(
            "If 'Others', please specify the Sales Line Name."
        )
        if not new_team:
            st.error("Please enter a valid Sales line.")
            return False, ""

        if not re.match(r"^[A-Z]{2}_.+$", new_team):
            st.error(
                "Sales line must follow the pattern: COUNTRYCODE_<Sales_Line_Name>"
            )
            return False, ""

        if new_team in st.session_state.options:
            st.warning("This Sales line already exists.")
            return False, ""

        return True, new_team  # New, valid team

    def select_region_and_country(self) -> None:
        """High-level orchestration wrapper shown in Streamlit UI."""
        st.markdown("<h3>Market & Brand Details</h3>", unsafe_allow_html=True)

        # 1. Set-up ---------------------------------------------------------------
        countries = self.get_master_country_list(ss.df_Team)

        # 2. Country selector -----------------------------------------------------
        st.session_state.country = self.select_country(countries)

        # 3. Sales team options for that country ----------------------------------
        self.update_sales_team_options(ss.df_Team, st.session_state.country)

    @staticmethod
    def num_of_days(date1, date2):
        """Calculate the number of days between two dates."""
        if date2 > date1:
            return (date2 - date1).days
        else:
            return (date1 - date2).days

    def set_cycle_dates(self):
        """Set the dates and length for the upcoming OCCP cycle."""
        today = datetime.datetime.now()
        next_year = today.year
        jan_1 = datetime.date(next_year, 1, 1)

        st.markdown(
            ("<div>" "<h3>Upcoming Cycle Details:</h3>" "</div>"),
            unsafe_allow_html=True,
        )
        # Input for Year and Month in YYYY/MM format
        cycle_start_input = st.text_input(
            "Select the Start Date (YYYY/MM)",
            value=f"{next_year}/01",
            help=(
                "Select the start date for the upcoming Omni Channel Call Plan (OCCP)  "
                "cycle in YYYY/MM format."
            ),
        )

        try:
            # Parse the input to get year and month
            year, month = map(int, cycle_start_input.split("/"))
            start_date = datetime.date(year, month, 1)
        except ValueError:
            st.error("Please enter the date in YYYY/MM format.")
            return

        cycle_length = st.number_input(
            "Select the Cycle Length (in months)",
            min_value=1,
            max_value=12,
            value=1,
            step=1,
            format="%d",
            help=(
                "Select the duration of the upcoming Omni Channel Call Plan (OCCP) cycle. "
                "Options are 1, 3, 4, or 6 months."
            ),
        )
        ss.cycle_start_date = start_date
        ss.cycle_length = cycle_length
        ss.cycle_end_date = start_date + relativedelta.relativedelta(
            months=cycle_length + 1, days=-1
        )
        ss.cycle_name = f"C{(start_date.month//cycle_length)+1} {start_date.year}"

        # Calculate the number of working days
        working_days = self.num_of_days(ss.cycle_start_date, ss.cycle_end_date)

        ss.working_days = st.number_input(
            "Enter the number of Working Days (in days)",
            min_value=1,
            max_value=working_days,
            value=1,
            step=1,
            format="%d",
            help=(
                "Enter the number of Working Days for the upcoming Omni Channel Call Plan (OCCP) cycle"
            ),
        )
        if ss.cycle_length > 1:
            ss.cycle_name = f"C{(start_date.month//cycle_length)+1} {start_date.year}"
        elif ss.cycle_length == 1:
            ss.cycle_name = f"C{(start_date.month//cycle_length)} {start_date.year}"

    def set_reference_dates(self):
        """Set the dates and length for the reference OCCP cycle."""
        today = datetime.datetime.now()
        prev_year = today.year - 1
        ref_jan_1 = datetime.date(prev_year, 1, 1)

        st.markdown(
            (
                "<link rel='stylesheet' href='https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css'>"
                "<div>"
                "<h3>Reference Cycle Details : <i class='fa fa-info-circle' aria-hidden='true' "
                "title='Reference cycle details are used to compare the Turing OCCP output for the upcoming cycle with the reference cycle OCCP. "
                "The OCCP output for the upcoming cycle is generated using the entire historical data.'></i></h3>"
                "</div>"
            ),
            unsafe_allow_html=True,
        )
        # Input for Year and Month in YYYY/MM format
        cycle_start_input = st.text_input(
            "Select the Start Date (YYYY/MM)",
            value=f"{prev_year}/01",
            help=(
                "Select the start date for the reference Omni Channel Call Plan (OCCP) cycle in YYYY/MM format."
            ),
        )

        try:
            year, month = map(int, cycle_start_input.split("/"))
            ref_start_date = datetime.date(year, month, 1)
        except ValueError:
            st.error("Please enter the date in YYYY/MM format.")
            return

        ref_length = st.number_input(
            "Select the Cycle Length (in months)",
            min_value=1,
            max_value=12,
            value=1,
            step=1,
            format="%d",
            help=(
                "Select the duration of the reference Omni Channel Call Plan (OCCP) cycle. "
                "Options are 1, 3, 4, or 6 months."
            ),
        )

        if ref_length != ss.cycle_length:
            st.warning(
                "**Note:**: Please verify, the reference cycle length and upcoming cycle length values are different."
            )

        ss.ref_start_date = ref_start_date
        ss.ref_length = ref_length
        ss.ref_end_date = ref_start_date + relativedelta.relativedelta(
            months=ref_length
        )

        ref_working_days = self.num_of_days(ss.ref_start_date, ss.ref_end_date)

        ss.ref_working_days = st.number_input(
            "Enter the number of Working Days (in days)",
            min_value=1,
            max_value=ref_working_days,
            value=1,
            step=1,
            format="%d",
            help=(
                "Enter the number of Working Days for the reference Omni Channel Call Plan (OCCP) cycle"
            ),
        )
        st.markdown("</div>", unsafe_allow_html=True)

    def send_email(
        self,
        subject: str,
        message: str,
        toaddrs: List,
        email_from: str,
        email_password: str,
        workbook,
        stmp_gateway,
        stmp_port,
    ) -> None:
        """Send an email with the Excel workbook attached."""
        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"] = email_from
        msg["To"] = ", ".join(toaddrs)
        part = MIMEText(message, "html")
        msg.attach(part)
        # ----#

        # json_file = MIMEApplication(json_object.encode('utf-8') , Name = 'data.json')
        # msg.attach(json_file)

        # ----#

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        attachment = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        attachment.set_payload(excel_buffer.read())
        encoders.encode_base64(attachment)
        brand = "_".join(ss.brands)
        attachment.add_header(
            "Content-Disposition",
            "attachment",
            filename=f"Business_Constraints_{ss.country}_{brand}_{ss.cycle_name}.xlsx",
        )
        msg.attach(attachment)

        # -----#
        mailserver = smtplib.SMTP(host=stmp_gateway, port=stmp_port)
        mailserver.starttls()
        # Use the EHLO command to identify the domain name of the
        # sending host to SMTP before you issue a MAIL FROM command.
        mailserver.ehlo()

        mailserver.login(user=email_from, password=email_password)
        mailserver.sendmail(email_from, toaddrs, msg.as_string())
        mailserver.close()
        st.toast("Email with Business Constraints has been sent", icon="🎉")
        st.success("OCCP Business constraints sent to OCCP POD team")

    def date_inputs(self):
        """Set both cycle and reference dates."""
        self.set_cycle_dates()
        self.set_reference_dates()

    def _brand_slider(self, brand, col, widget_id, disabled=False, help_text=""):
        """Helper to create a slider for brand distribution."""
        return col.slider(
            f"What percentage of the OCCP should be allocated to {brand}?",
            1,
            100,
            ss[f"brand_ratio_{brand}"],
            key=next(widget_id),
            step=1,
            disabled=disabled,
            help=help_text,
        )

    def show_header(self) -> None:
        """Render title + tooltip."""
        st.markdown(
            """
            <div class="tooltip">
                <h5>Brand Distribution:</h5>
                <span class="tooltiptext">
                    What percentage of the upcoming Omni Channel Call Plan (OCCP)
                    should be allocated for each selected brand.
                </span>
            </div>
            """,
            unsafe_allow_html=True,
        )

    def ensure_default_ratios(self, brands: list[str]) -> None:
        """Add a default even split into session_state if missing."""
        default = 100 // len(brands) if brands else 0
        for brand in brands:
            st.session_state.setdefault(f"brand_ratio_{brand}", default)

    def render_adjustable_sliders(
        self,
        brands: list[str],
        widget_id: count,
        col1,
        col2,
    ) -> None:
        """Draw sliders for all brands except the last and keep totals ≤ 100."""
        remaining = 100
        for i, brand in enumerate(brands[:-1]):
            col = col1 if i % 2 == 0 else col2

            # Maximum that still lets the rest reach 100
            others = sum(
                st.session_state[f"brand_ratio_{b}"] for b in brands[:-1] if b != brand
            )

            st.session_state[f"brand_ratio_{brand}"] = self._brand_slider(
                brand, col, widget_id
            )

            if st.session_state[f"brand_ratio_{brand}"] == 0:
                st.warning(
                    f"Warning: The distribution for {brand} is set to 0%. "
                    "Please adjust the value."
                )

    def render_last_brand_slider(
        self,
        brands: list[str],
        widget_id: count,
        col1,
        col2,
    ) -> None:
        """Auto-calculate last brand and display as disabled slider."""
        last = brands[-1]
        st.session_state[f"brand_ratio_{last}"] = max(
            0,
            100 - sum(st.session_state[f"brand_ratio_{b}"] for b in brands[:-1]),
        )

        if st.session_state[f"brand_ratio_{last}"] == 0:
            st.warning(
                f"Warning: The distribution for {last} is set to 0%. "
                "Please adjust the value."
            )

        col = col1 if (len(brands) - 1) % 2 == 0 else col2
        self._brand_slider(
            last,
            col,
            widget_id,
            disabled=True,
            help_text="Choose the brand distribution (in %) for multibrand OCCP interactions",
        )

    def brand_distribution(self) -> None:
        """Top-level orchestrator (very small & linear)."""
        self.show_header()

        col1, col2 = st.columns(2)
        widget_id = count(1)
        brands = ss.brands

        if not brands:
            st.info("No brands selected.")
            return

        self.ensure_default_ratios(brands)

        if len(brands) > 1:
            self.render_adjustable_sliders(brands, widget_id, col1, col2)
            self.render_last_brand_slider(brands, widget_id, col1, col2)
        else:  # Only a single brand: force 100 %
            brand = brands[0]
            st.session_state[f"brand_ratio_{brand}"] = 100
            self._brand_slider(brand, col1, widget_id, disabled=True)

        ss["brand_ratio"] = {
            brand: st.session_state[f"brand_ratio_{brand}"] for brand in brands
        }

    def select_brands_and_channels(self):
        """Handle brand and channel selection in the UI."""
        brand_name_mapping = {"SULIQUA": "SOLIQUA"}
        brand_list = list(ss.df_Brand["GLOBAL_BRAND"].unique())
        dupi_indication = list(
            ss.df_Brand[ss.df_Brand["GLOBAL_BRAND"] == "DUPIXENT"][
                "INDICATION_NAME"
            ].unique()
        )
        brand_to_remove = [
            "DUPIXENT",
            "CHRONIC RHINOSINUSITIS WITH NASAL POLYPS (CRSWNP)",
        ]
        brand_list = [brand for brand in brand_list if brand not in brand_to_remove]
        dupi_indication = [
            indi for indi in dupi_indication if indi not in brand_to_remove
        ]
        dupi_indication = [f"DUPIXENT {indi}" for indi in dupi_indication]
        brand_list = brand_list + dupi_indication
        sorted_brands = sorted(
            [brand for brand in brand_list if isinstance(brand, str)],
            key=lambda x: x.lower(),
        )
        sorted_brands.insert(0, CHOOSE_O)

        ss.option_selected = st.radio(
            "Choose the OCCP type",
            ["Monobrand", "Multibrand"],
            horizontal=True,
            help="Specify whether the Omni Channel Call Plan (OCCP) is for a single brand (Monobrand) or multiple brands (Multibrand).",
        )

        if ss.option_selected == "Monobrand":
            self._handle_monobrand_selection(sorted_brands)
        else:
            self._handle_multibrand_selection(sorted_brands)

    def _handle_monobrand_selection(self, sorted_brands):
        """Handle UI and logic for Monobrand selection."""
        selected_brand = st.selectbox(
            "Select the brand name",
            sorted_brands,
            label_visibility="visible",
            help="Choose a brand for the upcoming Omni Channel Call Plan (OCCP) cycle. If the brand name is not listed, contact the OCCP Support team.",
        )
        ss.brands = [selected_brand]
        ss.brand_id = ["BRAND1"]
        ss.brand_map = {selected_brand: ss.brand_id[0]}
        if selected_brand == CHOOSE_O:
            st.write("Please select a valid option.")

    def _handle_multibrand_selection(self, sorted_brands):
        """Handle UI and logic for Multibrand selection."""
        ss.brands = st.multiselect(
            "Select the brand names",
            sorted_brands[1:],
            label_visibility="visible",
            help="Choose brands for the upcoming Omni Channel Call Plan (OCCP) cycle. If the brand name is not listed, contact the OCCP Support team.",
        )
        ss.brands = sorted(ss.brands)
        ss.brand_id = [f"BRAND{i+1}" for i in range(len(ss.brands))]
        ss.brand_map = {
            brand: brand_id for brand, brand_id in zip(ss.brands, ss.brand_id)
        }

        if len(ss.brands) == 1 and not any(
            brand.startswith("DUPIXENT") for brand in ss.brands
        ):
            st.warning("Please select more than 1 brand")

        ss.specialties = {}
        if ss.brands:
            brands_str = " and ".join(ss.brands)
            specialty = st.text_input(
                f"Please specify Specialties for {brands_str} that can be promoted together."
            )
            st.warning(
                "**Note:** In case of multiple specialties, please input the text as comma-separated values."
            )
            if specialty:
                ss.specialties[brands_str] = specialty

    def set_channel_capacity(self, widget_id):
        """Set channel capacity sliders in the UI."""
        # Ensure monobrand_channel and multibrand_channel are lists
        if not isinstance(ss.monobrand_channel, list):
            ss.monobrand_channel = []
        if not isinstance(ss.multibrand_channel, list):
            ss.multibrand_channel = []

        # Combine unique channels from both monobrand and multibrand while preserving order
        self.unique_channels = list(
            dict.fromkeys(ss.monobrand_channel + ss.multibrand_channel)
        )

        if ss.is_non_prescriber == "No":
            hcp_type = "prescribers only"
        else:
            hcp_type = "prescribers and non-prescribers combined"
        i = 0
        if ss.is_non_prescriber == "Yes":
            st.markdown(
                """
            <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <div>
                <h6> Prescribers and Non-Prescribers  <i class="fa fa-info-circle" aria-hidden="true" title="It is the average number of interactions each prescriber and non-prescriber combined would be having per Rep per Day per Channel "></i></h6>
            </div>
            """,
                unsafe_allow_html=True,
            )

        cols = st.columns(2)
        for channel in self.unique_channels:
            with cols[i]:
                st.markdown(
                    f'<div style="height: 25px;">Average interactions per REP per day for {channel} for {hcp_type}.</div>',
                    unsafe_allow_html=True,
                )
                ss[f"channel_capacity_{channel}"] = st.number_input(
                    label=" ",
                    min_value=0.0,
                    value=0.0,
                    step=0.1,
                    key=f"{next(widget_id)}_hist",
                    format="%.1f",
                    help=f"Enter the average number of interactions a REP has per day for the {channel} ({hcp_type}).",
                )
            if i == 1:
                i = 0
            else:
                i = i + 1

    def configure_rep_capacity_constraints(self, obj):
        """Configure representative capacity constraints."""
        widget_id = count(1)
        st.markdown(
            """
            <div class="tooltip">
                <h5>Channel Constraints for Upcoming Cycle:</h5>
                <span class="tooltiptext">Choose the channels for the upcoming Omni Channel Call Plan (OCCP) cycle for the selected brand. Options include channels such as Remote and Face to Face.</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        ss.brand_ratio = {}
        # Define channels
        channels = [
            "F2F",
            "Remote",
            "Phone",
            "Meetings",
            "Virtual Meetings",
            "Whatsapp/Instant Message",
            "RTE-Open",
            "RTE-Sent",
        ]

        # TODO: Temporary adjustments in the code for RTE-Open and RTE-Sent
        # channels.extend(["RTE-Open","RTE-Sent"])
        # channels.remove("RTE/Approved Email")

        ss.monobrand_channel = []
        ss.multibrand_channel = []
        brand_channels = []
        if ss.option_selected == "Monobrand":
            selected_brand = ss.brands[0]
            ss.monobrand_channel = st.multiselect(
                "Select the OCCP channels in scope",
                channels,
                help="Choose the channels for upcoming OCCP Cycle for Monobrand OCCP interactions.",
            )

        elif ss.option_selected == "Multibrand":

            if len(ss.brands) >= 2:
                ss.monobrand_channel = st.multiselect(
                    "Select the OCCP channels in scope",
                    channels,
                    help="Choose the channels for the upcoming Omni Channel Call Plan (OCCP) cycle for Monobrand OCCP interactions.",
                )
                ss.multibrand_channel = st.multiselect(
                    "Select which Channel(s) can be Multibrand interactions?",
                    ss.monobrand_channel,
                    help="Choose the channels for the upcoming Omni Channel Call Plan (OCCP) cycle for Multibrand interactions.",
                )

        # Check if "RTE/Approved Email" is selected and display the E-Consent checkbox

        ss.econsent_rte = True
        # TODO Code adjustments for RTE-Open/Sent
        is_any_rte_present = any(
            _channel in ss.monobrand_channel or _channel in ss.multibrand_channel
            for _channel in ["RTE-Open", "RTE-Sent"]
        )
        is_both_rte_present = all(
            _channel in ss.monobrand_channel or _channel in ss.multibrand_channel
            for _channel in ["RTE-Open", "RTE-Sent"]
        )

        if is_both_rte_present:
            st.error("Please select either RTE-Open or RTE-Sent")

        if is_any_rte_present:
            econsent_checkbox_value = st.checkbox(
                "E-consent Required for Rep Triggered Email (RTE)", value=True
            )
            if not econsent_checkbox_value:
                ss.econsent_rte = False

        if ss.option_selected == "Multibrand":
            self.brand_distribution()

        st.markdown(
            """
            <div class="tooltip">
                <h5>Rep Capacity Constraints for Upcoming Cycle:</h5>
                <span class="tooltiptext">Enter the average number of interactions a REP has per day for the channel. For example, a REP can have the capacity to hold one face to face meeting with health care professional per day on average, and about 5 remote meetings</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        ss.is_non_prescriber = st.radio(
            "Is Non-prescriber included in the Target list?",
            ("No", "Yes"),
            horizontal=True,
        )

      
        ss.non_prescribers_priority = "Low"
        if ss.is_non_prescriber == "Yes":
            st.info(
                "**Note:** All non-prescribers and prescribers need to be identified by the "
                "country when sharing the target universe in the HCP-Rep mapping file."
            )

            ss.non_prescribers_priority = st.selectbox(
                "Select the level of Non-Prescribers priority",
                ["Low", "Medium", "High"],
            )

        self.set_channel_capacity(widget_id)

    def configure_hcp_capacity_constraints(self):
        """Configure HCP capacity constraints for the cycle."""
        widget_id = count(1)
        st.markdown(
            """
            <div class="tooltip">
                <h5>HCP Capacity Constraints for Upcoming Cycle:</h5>
                <span class="tooltiptext">Provide the minimum and maximum interaction limits based on reference data for the channel. For example, if HCP has had 1 interaction in the reference cycle/ segment level then the minimum and maximum interaction should be around the permissible range that seems feasible, like 1 and 4</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if ss.is_non_prescriber == "Yes":
            st.markdown(
                """
                <div class="tooltip">
                    <h5>(A) Inputs for Prescribers</h5>
                </div>
                """,
                unsafe_allow_html=True,
            )

        toggle_option = st.radio(
            "Select from the options below at which level the envelope matrix should be provided:",
            (HISTORICAL_INTERACTION_LEVEL, HCP_SEGMENT_LEVEL),
            horizontal=True,
        )

        ss.selected_envelope_toggle = toggle_option

        final_hcp_bounds = pd.DataFrame(
            columns=[
                "CHANNEL",
                "REFERENCE_CYCLE_ACTUAL",
                "MIN_VALUE",
                "MAX_VALUE",
            ]
        )
        final_hcp_segments = pd.DataFrame(
            columns=["CHANNEL", "SEGMENT", "MIN_VALUE", "MAX_VALUE"]
        )
        processed_channels = set()

        def process_channel(channel, col, is_segment=False, brand=None):
            """Helper to process a single channel's constraints."""

            if f"{channel} | {brand}" in processed_channels:
                return
            processed_channels.add(f"{channel} | {brand}")
            channel_upper = channel.upper().replace(" ", "_").replace("REP_", "")
            full_key = f"REP_{channel_upper}_{brand}_{next(widget_id)}"
            if is_segment:
                col.markdown(
                    f'<div style="height: 40px;">For {channel}, input the min/max interactions for segments.</div><br>',
                    unsafe_allow_html=True,
                )
                historical_df = pd.DataFrame(
                    [
                        {"Segment": seg, "Min": 0, "Max": 0}
                        for seg in ["A", "B", "C", "D"]
                    ]
                )
                edited_df = col.data_editor(
                    historical_df, num_rows="dynamic", key=full_key
                )
                hcp_bounds = edited_df.copy()
                hcp_bounds.columns = ["SEGMENT", "MIN_VALUE", "MAX_VALUE"]
                hcp_bounds["CHANNEL"] = f"{channel}"
                hcp_bounds["BRAND"] = f"{brand.upper()}"
                nonlocal final_hcp_segments
                final_hcp_segments = pd.concat(
                    [final_hcp_segments, hcp_bounds], ignore_index=True
                )
            else:
                col.markdown(
                    f'<div style="height: 40px;">For {channel}, input the min/max interactions limits.</div><br>',
                    unsafe_allow_html=True,
                )
                historical_df = pd.DataFrame(
                    [
                        {"Reference Cycle Actual": i, "Min": 0, "Max": 0}
                        for i in range(5)
                    ]
                )
                edited_df = col.data_editor(
                    historical_df, num_rows="dynamic", key=full_key
                )
                hcp_bounds = edited_df.copy()
                hcp_bounds.columns = [
                    "REFERENCE_CYCLE_ACTUAL",
                    "MIN_VALUE",
                    "MAX_VALUE",
                ]
                hcp_bounds["CHANNEL"] = channel_upper
                nonlocal final_hcp_bounds
                final_hcp_bounds = pd.concat(
                    [final_hcp_bounds, hcp_bounds], ignore_index=True
                )

        self._process_channels_by_column(
            ss.monobrand_channel, process_channel, toggle_option
        )

        if toggle_option == HISTORICAL_INTERACTION_LEVEL:
            ss.final_hcp_bounds_dict = self._group_hcp_bounds(
                final_hcp_bounds, "REFERENCE_CYCLE_ACTUAL"
            )
            return final_hcp_bounds
        else:
            ss.final_hcp_segments_dict = self._group_hcp_bounds(
                final_hcp_segments, "SEGMENT"
            )
            ss.e_consent_df, ss.rep_occp_df, ss.characterstics_df = (
                self.fetch_validate_data()
            )
            channel_utilization_status(ss, unique_channels=self.unique_channels)

            return final_hcp_segments

    # Adding Non-Precribers Constraints
    def non_prescribers_constraints(self):

        channels = ss.monobrand_channel
        channels_upper = [
            channel.upper().replace(" ", "_").replace("REP_", "")
            for channel in channels
        ]
        non_prescribers_constraints_df = pd.DataFrame(
            [{"Channel": channel, "Min": 0, "Max": 0} for channel in channels_upper]
        )
        if ss.is_non_prescriber == "No":
            return non_prescribers_constraints_df
        st.markdown(
            """
            <div class="tooltip">
                <h5>(B) Inputs for Non-Prescribers</h5>
            </div>
            """,
            unsafe_allow_html=True,
        )
        non_prescriber_input_msg = (
            "For non-prescribers, input the min/max interactions limits"
        )
        if ss.option_selected != "Monobrand":
            non_prescriber_input_msg += " (For All Brands)"
        st.write(f"{non_prescriber_input_msg}.")

        edited_non_prescribers_constraints_df = st.data_editor(
            non_prescribers_constraints_df, num_rows="dynamic"
        )
        final_edited_non_prescribers_constraints_df = (
            edited_non_prescribers_constraints_df.copy()
        )
        return final_edited_non_prescribers_constraints_df

    def _process_channels_by_column(self, channels, process_channel, toggle_option):
        """Helper to process channels across columns for UI."""
        is_segment=(toggle_option == HCP_SEGMENT_LEVEL)
        if is_segment:
            for brand in ss.brands:

                if len(ss.brands) > 1:
                    st.markdown(
                        f"""<ul>
                        <li>
                        <h5>Input the min/max interactions for {brand}</h5>
                        </li>
                        </ul>""",
                        unsafe_allow_html=True,
                    )

                cols = st.columns(3)
                for idx, channel in enumerate(channels):
                    col = cols[idx % 3]
                    process_channel(
                        channel,
                        col,
                        is_segment=(toggle_option == HCP_SEGMENT_LEVEL),
                        brand=brand,
                    )
        else:
            cols = st.columns(3)
            for idx, channel in enumerate(channels):
                col = cols[idx % 3]
                process_channel(
                    channel, col, is_segment=(toggle_option == HCP_SEGMENT_LEVEL)
                )

    def _group_hcp_bounds(self, df, group_col):
        """Helper to group HCP bounds into dict."""
        grouped_dict = {}
        if 'BRAND' in df.columns:
            df['CHANNEL_BRAND'] = df['CHANNEL']+" | "+df['BRAND']
            grouped = df.groupby(["CHANNEL_BRAND"])
        else:
            grouped = df.groupby(["CHANNEL"])
        for channel, group in grouped:
            if channel not in grouped_dict:
                grouped_dict[channel] = {}
            for _, row in group.iterrows():
                value = row[group_col]
                if isinstance(value, float):
                    key = str(int(value))
                else:
                    key = str(value)
                grouped_dict[channel][key] = [
                    int(row["MIN_VALUE"]),
                    int(row["MAX_VALUE"]),
                ]
        return grouped_dict

    def format_msg(self):
        """Format the subject and body for the business constraints email."""
        if len(ss["brands"]) == 1:
            subject = (
                "Business Constraints Submission for Review | "
                + str(ss["country"])
                + ", "
                + str(ss["brands"][0])
            )
            email_body = """<html>
                <head></head>
                <body>
                <h5><span lang=EN style='font-family:"Noto Sans",sans-serif;mso-fareast-font-family:"Times New Roman";color:#000000;mso-ansi-language:EN;font-weight:normal'>
                Hi,<br><br>
                We would like to inform you that a user has submitted the business constraints for the following details:
                <ul>
                <li>Country : <b>{}</b> </li> 
                <li>Brand : <b>{}</b> </li>
                <li>Sales Line : <b>{}</b> </li>
                <li>Cycle Name : <b>{}</b> </li>
                </ul>
                <br>
                Please find the attached Excel file containing the detailed constraints submitted through the OCCP Business Constraints Tool for your review.
                <br><br>
                Thank you for your attention to this matter.
                <br><br>
                Regards,
                <br>
                OCCP Business Constraints Tool
                <o:p></o:p></span></h5></body></html>""".format(
                ss["country"],
                ss["brands"][0],
                ss["sales_team"],
                ss["cycle_name"],
            )
        else:
            brands = ", ".join(ss["brands"])
            subject = (
                "Business Constraints Submission for Review | "
                + str(ss["country"])
                + " | "
                + brands
            )
            email_body = """<html>
                <head></head>
                <body>
                <h5><span lang=EN style='font-family:"Noto Sans",sans-serif;mso-fareast-font-family:"Times New Roman";color:#000000;mso-ansi-language:EN;font-weight:normal'>
                Hi,<br><br>
                We would like to inform you that a user has submitted the business constraints for the following details:
                <ul>
                <li>Country : <b>{}</b> </li> 
                <li>Brands : <b>{}</b> </li>
                <li>Sales Line : <b>{}</b> </li>
                <li>Cycle Name : <b>{}</b> </li>
                </ul>
                <br>
                Please find the attached Excel file containing the detailed constraints submitted through the OCCP Business Constraints Tool for your review.
                <br><br>
                Thank you for your attention to this matter.
                <br><br>
                Regards,
                <br>
                OCCP Business Constraints Tool
                <o:p></o:p></span></h5></body></html>""".format(
                ss["country"], brands, ss["sales_team"], ss["cycle_name"]
            )

        return subject, email_body

    def get_output_config(self):
        """Generate the output configuration dictionary for OCCP constraints."""
        interaction_channels, combined_interactions_dict, channel_rep_col = (
            self._build_interaction_channels()
        )
        constraints = self._build_constraints(
            interaction_channels, combined_interactions_dict
        )
        months_to_optimize, actual_months = self._get_months()
        feat_lst = interaction_channels + self._default_features()
        channel_cols_all = interaction_channels + channel_rep_col

        selected_country_code = self._get_country_code(ss.country.upper())
        multi_brand_sel = ss.option_selected != "Monobrand"
    
        if ss.is_non_prescriber=='Yes':
            non_prescribers_bounds = self._build_non_prescriber_constraints()
            constraints["NON_PRESCRIBERS_ENVELOPE_RULES"] = non_prescribers_bounds

        result = {
            "country_code": selected_country_code,
            "country": ss.country.upper(),
            "brand": ss.brands,
            "all_channel_cols": interaction_channels,
            "interaction_channels": interaction_channels,
            "channel_cols": interaction_channels,
            "channel_cols_all": channel_cols_all,
            "feature_list": feat_lst,
            "multibrand_data": multi_brand_sel,
            "max_country": True,
            "occp_length": ss.cycle_length,
            "months_to_optimize": months_to_optimize,
            "actual_date_range": actual_months,
            "CYCLE_NAME": ss.cycle_name.replace(" ", "_"),
            "CYCLE_START_DATE": str(ss.cycle_start_date),
            "CYCLE_END_DATE": str(ss.cycle_end_date),
            "e_consent": ss.econsent_rte,
            "combined_interactions_dict": combined_interactions_dict,
            "n_months": months_to_optimize,
            "channel_brand_dict": combined_interactions_dict,
            "capacity_constraints": self._build_capacity_constraints(channel_rep_col),
            "constraints": constraints,
            "is_non_prescriber": ss.is_non_prescriber,
            "non_prescribers_priority": ss.non_prescribers_priority,
        }
        return result

    def _build_interaction_channels(self):
        """Helper to build interaction channels and related dicts."""
        interaction_channels = []
        combined_interactions_dict = {}
        channel_rep_col = []
        channel_map = self._channel_map()
        ss.brand_map = {brand: f'BRAND{i+1}' for i, brand in enumerate(ss.brand_map)}   
        for channel in ss.monobrand_channel:
            channel = channel.upper()
            mapped_channel = channel_map.get(channel, channel).replace(" ", "_")
            channel_key = f"REP_{mapped_channel}"
            channel_rep_col.append(channel_key)
            combined_interactions_dict[channel_key] = []
            for brand in ss.brands:
                if brand in ss.brand_map:
                    interaction = f"{channel_key}_{ss.brand_map[brand]}"
                    interaction_channels.append(interaction)
                    combined_interactions_dict[channel_key].append(interaction)
            if any(channel.lower() == ch.lower() for ch in ss.multibrand_channel):
                for combo in self._brand_combinations():
                    if all(b in ss.brand_map for b in combo):
                        combo_key = "_AND_".join([ss.brand_map[b] for b in combo])
                        interaction = f"{channel_key}_{combo_key}"
                        interaction_channels.append(interaction)
                        combined_interactions_dict[channel_key].append(interaction)
        return interaction_channels, combined_interactions_dict, channel_rep_col

    def _build_constraints(self, interaction_channels, combined_interactions_dict):
        """Helper to build constraints dict."""
        if ss.selected_envelope_toggle == HISTORICAL_INTERACTION_LEVEL:
            return {
                "ENVELOPE_RULES": self._transform_hcp_bounds(
                    ss.final_hcp_bounds_dict, interaction_channels
                )
            }
        else:
            return {
                "ENVELOPE_RULES": self._transform_hcp_segments(
                    ss.final_hcp_segments_dict,
                    is_segment=True,
                )
            }

    def _build_non_prescriber_constraints(self):
        non_prescribers_bounds_dict = {
            row["Channel"]: [row["Min"], row["Max"]]
            for _, row in self.non_prescribers_constraints_df.iterrows()
        }
        channel_map = self._channel_map()
        non_prescribers_bounds = {
            f"REP_{channel_map.get(k,k)}": v
            for k, v in non_prescribers_bounds_dict.items()
        }
        return non_prescribers_bounds

    def _transform_hcp_segments(self, hcp_bounds_dict, is_segment=False):
        """Helper to transform HCP segment bounds for output config"""
        transformed = {}
        channel_map = self._channel_map()
        brand_grouped = defaultdict(dict)

        for key, segments in hcp_bounds_dict.items():
            channel, brand_name = key.split(" | ", 1)
            channel = channel.upper()
            brand_grouped[brand_name][
                f"REP_{channel_map.get(channel,channel)}"
            ] = segments

        brand_grouped = dict(sorted(brand_grouped.items()))
        transformed = dict(
            chain.from_iterable(
                [(f"{channel}_BRAND{i}", segment) for channel, segment in inner.items()]
                for i, (_, inner) in enumerate(brand_grouped.items(), start=1)
            )
        )
        return transformed

    def _transform_hcp_bounds(
        self, hcp_bounds_dict, interaction_channels, is_segment=False
    ):
        """Helper to transform HCP bounds for output config."""
        transformed = {}
        channel_map = self._channel_map()
        normalized_multibrand_channels = [
            ch.lower().replace("_", " ").strip() for ch in ss.multibrand_channel
        ]
        for channel_tuple, brand_bounds in hcp_bounds_dict.items():
            channel_raw = str(channel_tuple).replace("_", " ").strip().lower()
            mapped_channel = channel_map.get(
                channel_raw.upper(), channel_raw.upper()
            ).replace(" ", "_")
            is_multibrand = channel_raw in normalized_multibrand_channels
            brand_combos = (
                self._brand_combinations()
                if is_multibrand
                else [(brand,) for brand in ss.brands]
            )
            for brand in ss.brands:
                if brand in ss.brand_map:
                    key = f"REP_{mapped_channel}_{ss.brand_map[brand]}"
                    transformed[key] = brand_bounds
            key = f"REP_{mapped_channel}"
            transformed[key] = brand_bounds
            for combo in brand_combos:
                if all(brand in ss.brand_map for brand in combo):
                    brand_key = "_AND_".join([ss.brand_map[brand] for brand in combo])
                    key = f"REP_{mapped_channel}_{brand_key}"
                    transformed[key] = brand_bounds
        return transformed

    def _brand_combinations(self):
        """Helper to get all brand combinations for multibrand."""
        return [
            combo
            for r in range(2, len(ss.brands) + 1)
            for combo in combinations(ss.brands, r)
        ]

    def _get_months(self):
        """Helper to get months to optimize and actual months."""
        months_to_optimize = [
            ss.cycle_start_date + pd.DateOffset(months=i)
            for i in range(ss.cycle_length)
        ]
        months_to_optimize = [
            month.strftime(date_format_str) for month in months_to_optimize
        ]
        actual_cycle_length_in_months = ss["ref_length"]
        actual_months = [
            ss["ref_start_date"] + pd.DateOffset(months=i)
            for i in range(actual_cycle_length_in_months)
        ]
        actual_months = [month.strftime(date_format_str) for month in actual_months]
        return months_to_optimize, actual_months

    def _default_features(self):
        """Helper for default features list."""
        return [
            "MONTH_1",
            "MONTH_2",
            "MONTH_3",
            "MONTH_4",
            "MONTH_5",
            "MONTH_6",
            "MONTH_7",
            "MONTH_8",
            "MONTH_9",
            "MONTH_10",
            "MONTH_11",
            "BRICK_SALES_AMOUNT_MA_3",
        ]

    def _channel_map(self):
        """Helper for channel mapping."""
        return {
            "F2F": "FACE_TO_FACE",
            "REMOTE": "REMOTE_MEETING",
            "WHATSAPP/INSTANT MESSAGE": "WHATSAPP",
            "RTE-OPEN": "TRIGGERED_EMAIL",
            "RTE-SENT": "TRIGGERED EMAIL",
        }

    def _get_country_code(self, selected_country):
        """Helper to get country code."""
        country_codes = {
            "AR": "ARGENTINA",
            "ES": "SPAIN",
            "KR": "SOUTH KOREA",
            "BR": "BRAZIL",
            "NL": "NETHERLANDS",
            "TR": "TURKEY",
            "JP": "JAPAN",
            "GC": "SAUDI ARABIA",
            "FR": "FRANCE",
            "IT": "ITALY",
            "BE": "BELGIUM",
            "DE": "GERMANY",
            "SE": "SWEDEN",
            "PL": "POLAND",
            "MX": "MEXICO",
            "AU": "AUSTRALIA",
            "CA": "CANADA",
            "CO": "COLOMBIA",
            "UAE": "UAE",
        }
        for code, country in country_codes.items():
            if country == selected_country:
                return code
        return ""

    def _build_capacity_constraints(self, channel_rep_col):
        channel_map = self._channel_map()
        daily_capacity_channels_dict = {}
        for rep_channel in channel_rep_col:
            canonical = rep_channel.replace("REP_", "").upper()  # FACE_TO_FACE
            raw_label = channel_map.get(canonical, canonical)
            key = f"channel_capacity_{raw_label}"
            value = float(ss.get(key, 0))
            daily_capacity_channels_dict[rep_channel] = int(
                (value * ss.working_days) / ss.cycle_length
            )

        return (
            {ss.sales_team: daily_capacity_channels_dict}
            if ss["sales_team"]
            else daily_capacity_channels_dict
        )

    def resolve_brands(self, list_brands):
        """
        Resolve the brand names to their corresponding codes.
        """
        resolved_brands = []
        for brand in list_brands:
            resolved_brands.append(get_brand_code(brand))
        return "_".join(resolved_brands)

    def submit_payload(self):
        """Submit the OCCP payload and send notification email."""
        with open(email_path, "r", encoding="utf-8") as f:
            email_config = yaml.safe_load(f)

        sel_cntry_code = self._get_country_code(ss.country.upper())

        if ENV == "PROD":
            url = "https://apps.factoryv2.p171649450587.aws-emea.sanofi.com/prod/turing-geneticalgorithm/trigger/run"
            # Base recipients from global 'to.emails'
            to_emails = email_config.get("to", {}).get("emails", []).copy()

            # Add 'to.gen.<country_code>.emails' if available
            country_specific_emails = (
                email_config.get("to", {})
                .get("gen", {})
                .get(sel_cntry_code, {})
                .get("emails", [])
            )
            to_emails.extend(
                email for email in country_specific_emails if email not in to_emails
            )
        else:
            url = "https://apps.factoryv2.p171649450587.aws-emea.sanofi.com/uat/turing-geneticalgorithm/trigger/run"
            to_emails = email_config.get("test", {}).get("emails", []).copy()

        result = self.get_output_config()

        if not result:
            st.error(
                "Error in generating the output configuration. Please check the inputs."
            )
            return

        data = {
            "gbu": "gen",
            "countrycode": result["country_code"],
            "brand": self.resolve_brands(result["brand"]),
            "subtype": "BRICK",
            "constraints": json.dumps(result),
        }
        
        if ENV != "DEV":
            payload_json = json.dumps(data)
            response = requests.post(
                url, data=payload_json, verify="/etc/ssl/certs/ca-certificates.crt"
            )

            if response.status_code == 200:
                st.success("Payload submitted successfully!")
                response_data = response.text
            else:
                st.error(
                    f"Failed to submit payload. Status code: {response.status_code}"
                )
                st.write(response.text)

        # Optionally add user email
        if self.user_email and self.user_email not in to_emails:
            to_emails.append(self.user_email)

        subject, body = self.format_msg()

        email_pwd = os.environ.get("EMAIL_PWD")

        if email_pwd:
            self.send_email(
                subject,
                body,
                to_emails,
                email_config["from"][0],
                email_pwd,
                self.wb,
                email_config["SMTP_GATEWAY"][0],
                email_config["SMTP_PORT"][0],
            )

    def calculate_business_constraints(self, final_hcp_bounds):
        """Calculate business constraints and return summary DataFrames."""
        sales_line = {"Input": f"{ss.sales_team}"}
        constant_features = {
            "Generate OCCP for cycle": f"{ss.cycle_start_date} - {ss.cycle_end_date}",
        }

        multibrand_flag = "Multibrand" if len(ss.brands) > 1 else "Monobrand"

        occp_context_features = {
            "OCCP Cycle Length": f"{ss.cycle_length} months",
            "Number of Working Days for upcoming cycle": f"{ss.working_days}",
            "Reference Cycle": f"{ss.ref_start_date} - {ss.ref_end_date}",
            "OCCP Channels": ", ".join(self.unique_channels),
            "Monobrand/Multibrand OCCP": multibrand_flag,
            "OCCP Brand(s)": ", ".join(ss.brands),
            "For Multibrand OCCP only - CHANNELS": (
                ", ".join(self.unique_channels)
                if multibrand_flag == "Multibrand"
                else ""
            ),
        }

        daily_capacity_channels_dict = {
            f"Avg Rep Capacity per day for {channel}": f"{float(ss[f'channel_capacity_{channel}'])}"
            for channel in self.unique_channels
        }

        envelope_matrix_rows = []

        channel_data = final_hcp_bounds
        if "REFERENCE_CYCLE_ACTUAL" in final_hcp_bounds.columns:
            for idx, row in channel_data.iterrows():
                envelope_matrix_rows.append(
                    {
                        "CHANNEL": row["CHANNEL"],
                        "REFERENCE_CYCLE_ACTUAL": row["REFERENCE_CYCLE_ACTUAL"],
                        "MIN_VALUE": row["MIN_VALUE"],
                        "MAX_VALUE": row["MAX_VALUE"],
                    }
                )
        elif "SEGMENT" in final_hcp_bounds.columns:
            for idx, row in channel_data.iterrows():
                envelope_matrix_rows.append(
                    {
                        "CHANNEL": row["CHANNEL"],
                        "BRAND": row["BRAND"],
                        "SEGMENT": row["SEGMENT"],
                        "MIN_VALUE": row["MIN_VALUE"],
                        "MAX_VALUE": row["MAX_VALUE"],
                    }
                )

        additional_comments = {
            "Do you need e-consent to be checked for RTE? (Yes/No)": ss.econsent_rte
        }

        hcp_constraints = {
            "Envelope matrix per HCP by channel": "Please refer to 'Envelope matrix - Input' tab"
        }

        combined_dict = {
            **sales_line,
            **constant_features,
            **occp_context_features,
            **daily_capacity_channels_dict,
            **hcp_constraints,
            **additional_comments,
        }

        data = [
            ("Type", sales_line),
            ("Cycle", constant_features),
            ("OCCP Context", occp_context_features),
            ("Sales Rep. Constraints", daily_capacity_channels_dict),
            (TYPE_LABEL, hcp_constraints),
            ("Additional Constraint", additional_comments),
        ]

        rows = []
        for source, dictionary in data:
            for key, value in dictionary.items():
                rows.append(
                    {
                        "Constraint type": source,
                        "Information": key,
                        "Values": value,
                    }
                )
        df = pd.DataFrame(rows)

        envelope_matrix_df = pd.DataFrame(envelope_matrix_rows)

        return df, envelope_matrix_df

    def create_excel(self, df, envelope_matrix_df, non_prescribers_envelop):
        """Create an Excel workbook from the constraints DataFrames."""

        # Create workbook and worksheet
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Business Constraints"

        # Add and format headers
        headers = ["Type", "Input", ss.sales_team]
        ws.append(headers)
        self._format_excel_headers(ws)

        # Add main business constraints data
        data = self._build_excel_data()
        for row in data:
            ws.append(row)
        self._merge_excel_cells(ws)
        self._format_excel_cells(ws)

        # Add envelope matrix section
        if "REFERENCE_CYCLE_ACTUAL" in envelope_matrix_df.columns:
            matrix_headers = ["Type", "REFERENCE_CYCLE_ACTUAL", "Min_Value", "Max_Value"]
            value_columns = ["REFERENCE_CYCLE_ACTUAL", "MIN_VALUE", "MAX_VALUE"]
            section_title = "Channel"
            group_by_col = "CHANNEL"
            type_label = TYPE_LABEL
        elif "SEGMENT" in envelope_matrix_df.columns:
            matrix_headers = ["Type", "BRAND", "SEGMENT", "Min_Value", "Max_Value"]
            value_columns = ["BRAND", "SEGMENT", "MIN_VALUE", "MAX_VALUE"]
            section_title = "Channel"
            group_by_col = "CHANNEL"
            type_label = TYPE_LABEL
        else:
            matrix_headers = value_columns = section_title = group_by_col = type_label = None

        if matrix_headers:
            self._add_envelope_matrix_generic(
                ws, envelope_matrix_df, section_title, group_by_col, matrix_headers, value_columns, type_label
            )

        # Add non-prescribers envelope matrix section
        self._add_non_prescribers_envelope_matrix(ws, non_prescribers_envelop)
        # Auto-adjust column widths
        self._auto_adjust_column_width(ws)

        # Save workbook to an in-memory buffer for further use (e.g., email)
        from io import BytesIO

        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        self.excel_bytes = output.getvalue()
        return output

    def _format_excel_headers(self, ws):
        """Helper to format Excel headers."""
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="7a00e6", end_color="7a00e6", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

    def _build_excel_data(self):
        """Helper to build data rows for Excel."""
        data = [
            ["Cycle", "Generate OCCP for cycle", f"{ss.cycle_name}"],
            [
                OCCP_CONTEXT,
                "OCCP Cycle Length",
                f"{ss.cycle_start_date.strftime('%b %Y')} - {ss.cycle_end_date.strftime('%b %Y')}",
            ],
            [
                OCCP_CONTEXT,
                "Number of Working Days for Upcoming cycle",
                f"{ss.working_days}",
            ],
            [
                OCCP_CONTEXT,
                "Reference Cycle",
                f"{ss.ref_start_date.strftime('%b %Y')} - {ss.ref_end_date.strftime('%b %Y')}",
            ],
            [
                OCCP_CONTEXT,
                "Number of Working Days for Reference cycle",
                f"{ss.ref_working_days}",
            ],
            [
                OCCP_CONTEXT,
                "Monobrand/Multibrand OCCP",
                "Multibrand" if len(ss.brands) > 1 else "Monobrand",
            ],
            [OCCP_CONTEXT, "OCCP Brand(s)", ", ".join(ss.brands)],
            [OCCP_CONTEXT, "OCCP Channels", ", ".join(ss.monobrand_channel)],
        ]
        if len(ss.brands) > 1:
            data.append(
                [
                    MULTIBRAND_OCCP,
                    "Select which channels can be multibrand interactions?",
                    ", ".join(ss.multibrand_channel),
                ]
            )
            if ss.specialties:
                specialties_str = ", ".join(
                    [
                        f"{brand_str} : {specialty}"
                        for brand_str, specialty in ss.specialties.items()
                    ]
                )
                data.append(
                    [
                        MULTIBRAND_OCCP,
                        "Specify Specialities that can be promoted together ?",
                        specialties_str,
                    ]
                )
            brand_distribution_str = ", ".join(
                [f"{ss['brand_ratio'][brand]}% {brand}" for brand in ss.brands]
            )
            data.append(
                [
                    MULTIBRAND_OCCP,
                    "Brand distribution",
                    brand_distribution_str,
                ]
            )
        data.append(
            [
                SALES_REP_CONSTRAINTS,
                "Avg Rep Capacity for All Channels",
                "{:.2f}".format(
                    sum(
                        float(ss.get(f"channel_capacity_{channel}", 0))
                        for channel in self.unique_channels
                    )
                ),
            ]
        )
        for channel in self.unique_channels:
            if ss.is_non_prescriber == "No":
                data.append(
                    [
                        SALES_REP_CONSTRAINTS,
                        f"Avg Rep Capacity per day for {channel}",
                        f"{float(ss.get(f'channel_capacity_{channel}', 0)):.2f}",
                    ]
                )
        for channel in self.unique_channels:
            if ss.is_non_prescriber == "Yes":
                data.append(
                    [
                        SALES_REP_CONSTRAINTS,
                        f"Avg Rep Capacity per day for {channel} (Prescriber and Non-Prescribers Combined)",
                        f"{float(ss.get(f'channel_capacity_{channel}', 0)):.2f}",
                    ]
                )

        if any(
            _channel in self.unique_channels for _channel in ["RTE-Open", "RTE-Sent"]
        ):
            consent_flag = "Yes" if ss.econsent_rte else "No"
            data.append(
                [
                    "Additional Constraints",
                    "eConsent required for RTE?",
                    consent_flag,
                ]
            )

        v_consent = "Yes" if ss.veeva_checkbox == True else "No"
        data.append(
            [
                "Veeva Align Format",
                "Do you require the OCCP output in Veeva Align format for sales representative review? (Yes/No)",
                v_consent,
            ]
        )
        return data

    def _merge_excel_cells(self, ws):
        """Helper to merge cells in Excel."""
        col = 1
        start_row = 2

        while start_row <= ws.max_row:
            current_value = ws.cell(start_row, col).value

            if not current_value:
                start_row += 1
                continue

            end_row = start_row + 1
            while (
                end_row <= ws.max_row and ws.cell(end_row, col).value == current_value
            ):
                end_row += 1

            if end_row - start_row > 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row - 1,
                    end_column=col,
                )
                ws.cell(start_row, col).alignment = Alignment(vertical="top")

            start_row = end_row

    def _format_excel_cells(self, ws):
        """Helper to format Excel cells."""
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )

    def _add_envelope_matrix_generic(
        self, ws, df, section_title, group_by_col, matrix_headers, value_columns, type_label
    ):
        """Generic helper to add envelope matrix sections to Excel."""
        if df.empty or group_by_col not in df.columns:
            return

        for group, group_df in df.groupby(group_by_col):
            # Blank row for spacing
            ws.append([""])
            # Section header row
            ws.append([f"{section_title}: {group}"])
            header_row_idx = ws.max_row
            for cell in ws[header_row_idx]:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="dcd0ff", end_color="dcd0ff", fill_type="solid")
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )
            ws.merge_cells(
                start_row=header_row_idx,
                start_column=1,
                end_row=header_row_idx,
                end_column=len(matrix_headers),
            )
            ws.cell(row=header_row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

            # Matrix headers
            ws.append(matrix_headers)
            matrix_header_row_idx = ws.max_row
            for cell in ws[matrix_header_row_idx]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="7a00e6", end_color="dcd0ff", fill_type="solid")
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )

            # Data rows
            start_row = ws.max_row + 1
            for _, row in group_df.iterrows():
                ws.append([""] + [row[col] for col in value_columns])
            end_row = ws.max_row

            # Merge the "Type" column for this group
            ws.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=end_row,
                end_column=1,
            )
            ws.cell(row=start_row, column=1).value = type_label
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

            # Add Excel table style for this matrix
            tab = Table(
                displayName=f"Table_{str(group).replace(' ', '_')}_{start_row}",
                ref=f"A{start_row}:D{end_row}",
            )
            style = TableStyleInfo(
                name="TableStyleLight12",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

    def _add_non_prescribers_envelope_matrix(self, ws, non_prescribers_envelop):
        """Helper to add envelope matrix to Excel."""
        if (
            non_prescribers_envelop.empty
            or "Channel" not in non_prescribers_envelop.columns
        ):
            return

        # Determine which columns to use for the matrix
        matrix_headers = [
            "Type",
            "CHANNEL",
            "Min_Value",
            "Max_Value",
        ]
        value_columns = ["Channel", "Min", "Max"]

        ws.append([""])
        # Blank row for spacing

        # Channel header row
        ws.append([f"NON-PRESCRIBERS ENVELOPE RULES"])
        header_row_idx = ws.max_row
        for cell in ws[header_row_idx]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(
                start_color="dcd0ff", end_color="dcd0ff", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
        ws.merge_cells(
            start_row=header_row_idx,
            start_column=1,
            end_row=header_row_idx,
            end_column=4,
        )
        ws.cell(row=header_row_idx, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Matrix headers
        ws.append(matrix_headers)
        matrix_header_row_idx = ws.max_row
        for cell in ws[matrix_header_row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="7a00e6", end_color="dcd0ff", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

            # Data rows
            start_row = ws.max_row + 1
        for _, row in non_prescribers_envelop.iterrows():
            ws.append([""] + [row[col] for col in value_columns])
        end_row = ws.max_row

        # Merge the "Type" column for this group
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=end_row,
            end_column=1,
        )
        ws.cell(row=start_row, column=1).value = "NON-PRESCRIBERS Constraints"
        ws.cell(row=start_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Add Excel table style for this matrix

        tab = Table(
            displayName=f"Table_Non_Prescribers_Envelope_Rules_{start_row}",
            ref=f"A{start_row}:D{end_row}",
        )
        style = TableStyleInfo(
            name="TableStyleLight12",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    def _auto_adjust_column_width(self, ws):
        """Helper to auto-adjust column widths."""
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    def push_output_tables_to_df(self, output_table_dict):
        for key, value in output_table_dict.items():
            write_pandas(self.output_snowflake_con, value, key)

    @st.dialog("Constraints Summary", width="large")
    def review_button(
        self,
        final_hcp_bounds,
        final_edited_non_prescribers_constraints_df,
        output_table_dict,
    ):
        """
        Show a dialog for reviewing constraints before submission.
        This method displays all relevant OCCP constraint details and handles user email input and submission.
        """
        self._render_market_details()
        self._render_cycle_details()
        self._render_reference_cycle_details()
        self._render_channel_details()
        self._render_rep_capacity_constraints()
        self._render_hcp_constraints(final_hcp_bounds)
        self._render_non_prescribers_details(
            final_edited_non_prescribers_constraints_df
        )
        self._render_email_section(output_table_dict)

    def _render_market_details(self):
        """Render market and brand details section."""
        self.render_header("1. Market & Brand Details:")
        details = {
            "Country": ss["country"],
            "Brands": ", ".join(ss["brands"]),
            "Sales Line": ss["sales_team"],
        }
        st.markdown("\n".join([f"* **{k}**: {v}" for k, v in details.items()]))
        if ss.option_selected == "Multibrand" and ss.specialties:
            for brands_str, specialty in ss.specialties.items():
                st.markdown(
                    f"* **Specialties for {brands_str} that can be promoted together**: {specialty}"
                )

    def _render_cycle_details(self):
        """Render upcoming cycle details section."""
        self.render_header("2. Upcoming Cycle Details")
        details = {
            "Cycle Name": ss["cycle_name"],
            "Cycle Length (in months)": str(ss["cycle_length"]),
            "Cycle Start Date": str(ss["cycle_start_date"]),
            "Working Days in Upcoming Cycle": str(ss["working_days"]),
        }
        st.markdown("\n".join([f"* **{k}**: {v}" for k, v in details.items()]))

    def _render_reference_cycle_details(self):
        """Render reference cycle details section."""
        self.render_header("3. Reference Cycle Details")
        details = {
            "Cycle Length (in months)": str(ss["ref_length"]),
            "Cycle Start Date": str(ss["ref_start_date"]),
            "Working Days in Reference Cycle": str(ss["ref_working_days"]),
        }
        st.markdown("\n".join([f"* **{k}**: {v}" for k, v in details.items()]))

    def _render_channel_details(self):
        """Render channel selection details section."""
        self.render_header("4. Channel Details")
        if ss.option_selected == "Monobrand":
            st.markdown(
                f"**OCCP channels in scope**: {', '.join(ss.monobrand_channel)}"
            )
        elif ss.option_selected == "Multibrand":
            st.markdown(
                f"* **OCCP channels in scope**: {', '.join(ss.monobrand_channel)}"
            )
            st.markdown(
                f"* **Channels for Multibrand OCCP only**: {', '.join(ss.multibrand_channel)}"
            )

    def _render_rep_capacity_constraints(self):
        """Render REP capacity constraints section."""
        self.render_header("5. REP Capacity Constraints")
        rep_capacity_dict = self._get_rep_capacity_dict()
        st.markdown(
            "\n".join([f"* **{k}**: {v}" for k, v in rep_capacity_dict.items()])
        )

    def _get_rep_capacity_dict(self):
        """Build REP capacity details dictionary."""
        rep_capacity_dict = {}
        for channel in self.unique_channels:
            if ss.is_non_prescriber == "Yes":
                rep_capacity_dict[
                    f"Avg {channel} interactions per day (Prescriber & Non-Prescriber combined)"
                ] = ss[f"channel_capacity_{channel}"]
            else:
                rep_capacity_dict[f"Avg {channel} interactions per day"] = ss[
                    f"channel_capacity_{channel}"
                ]

        return rep_capacity_dict

    def _render_hcp_constraints(self, final_hcp_bounds):
        """Render HCP constraints tables for each channel."""
        
        if "REFERENCE_CYCLE_ACTUAL" in final_hcp_bounds.columns:
            for group, subset in final_hcp_bounds.groupby("CHANNEL"):
                self.render_header(f"6. HCP Constraints for {group} ")
                st.markdown(
                    subset.to_html(
                        index=False,
                        columns=["REFERENCE_CYCLE_ACTUAL", "MIN_VALUE", "MAX_VALUE"],
                    ),
                    unsafe_allow_html=True,
                )
        elif "SEGMENT" in final_hcp_bounds.columns:
            for group, subset in final_hcp_bounds.groupby(["CHANNEL", "BRAND"]):
                channel, brand = group 
                self.render_header(f"6. HCP Constraints for {channel} | Brand: {brand}")
                st.markdown(
                    subset.to_html(
                        index=False,
                        columns=["SEGMENT", "MIN_VALUE", "MAX_VALUE"],
                    ),
                    unsafe_allow_html=True,
                )

                # # IMP: Need to fix Rep Utilization Logic
                # st.markdown("#### Details:")
                # message = ss.val_results[channel]['message']
                # for msg in message:
                #     st.markdown(f"- {msg}")

    def _render_non_prescribers_details(
        self, final_edited_non_prescribers_constraints_df
    ):
        """Render REP capacity constraints section."""

        self.render_header("7. Non-Prescribers Details")
        st.markdown(
            f"* **Is Non-prescriber included in the Target list**: {ss.is_non_prescriber}"
        )
        if ss.is_non_prescriber == "Yes":
            # st.markdown(
            #     f"* **Consent for Prescribers & Non-Prescriber Identefication by Country**: {ss.non_prescribers_consent_value}"
            # )
            st.markdown(
                f"* **Non-Prescribers priority**: {ss.non_prescribers_priority}"
            )

            self.render_header("8. Non-Prescribers Constraints")
            self.non_prescribers_constraints_df = (
                final_edited_non_prescribers_constraints_df
            )
            non_prescribers_html_table = self.non_prescribers_constraints_df.rename(
                columns={"Channel": "CHANNEL", "Min": "MIN_VALUE", "Max": "MAX_VALUE"}
            ).to_html(index=False)
            st.markdown(
                non_prescribers_html_table,
                unsafe_allow_html=True,
            )

    def _render_email_section(self, output_table_dict):
        """Render the email input and submission section."""
        columns = st.columns((1, 0.75, 1))
        with columns[0]:
            st.warning(
                "**Note:** In case of multiple email recipients, please input the email ids as comma-separated values."
            )
            self.user_email = st.text_input(
                "Provide your email address to receive a copy of the business constraints:"
            )
            if not self.user_email:
                st.warning("Provide your email address")
                return
            if not self.user_email.endswith("@sanofi.com"):
                st.warning("Please provide a valid sanofi email address.")
                return
            if st.button("Submit Business Constraints"):
                with st.spinner("Submitting..."):
                    self.push_output_tables_to_df(output_table_dict)
                    self.submit_payload()
            else:
                st.warning(
                    "Click on this button to share the Business Constraints with the OCCP team"
                )

    def render_header(self, header_text):
        """Render a colored section header in the review dialog."""
        st.markdown(
            f'<h2 style="color: #7a00e6;">{header_text}</h2>',
            unsafe_allow_html=True,
        )


def main():
    """Main entry point for the Streamlit OCCP tool."""
    obj = OCCPTool()
    obj.initialize_session_state()
    obj.configure_page(obj)
    with st.sidebar:
        image_path = os.path.join(SCRIPT_DIR, "./utils/turing.PNG")
        st.image(image_path)
        obj.select_region_and_country()
        obj.select_brands_and_channels()
        obj.set_cycle_dates()
        obj.set_reference_dates()

    if all(
        [
            (ss["working_days"] > 1),
            (ss["ref_working_days"] > 1),
            (ss.team_flag),
            (ss.sales_team != CHOOSE_O),
            (CHOOSE_O not in ss.brands),
            (ss.brands),
            (
                (ss.option_selected == "Monobrand")
                or (ss.option_selected != "Monobrand" and len(ss.brands) > 1)
            ),
        ]
    ):
        obj.configure_rep_capacity_constraints(obj)
        final_hcp_bounds = obj.configure_hcp_capacity_constraints()
        df, envelope_matrix_df = obj.calculate_business_constraints(final_hcp_bounds)
        final_edited_non_prescribers_constraints_df = obj.non_prescribers_constraints()
        df_channel = obj.fetch_channel_data_from_snowflake()
        df_master, df_time_dim = obj.fetch_master_and_time_dim_data_from_snowflake()

        df_team, df_brand = obj.fetch_team_data_from_snowflake()

        # Check if all required fields are filled
        required_fields_filled = (
            ss["country"]
            and ss["brands"]
            and ss["cycle_name"]
            and ss["sales_team"]
            and ss["cycle_start_date"]
            and ss["working_days"]
            and all(
                ss.get(f"channel_capacity_{channel}") for channel in obj.unique_channels
            )
        )


        st.markdown(
            """
            <div class="tooltip">
                <h5>Veeva Align Format:</h5>
                </div>
            """,
            unsafe_allow_html=True,
        )
        ss.veeva_checkbox = False
        veeva_checkbox_value = st.radio(
            "Do you require the OCCP output in Veeva Align format for review by Sales Representative?",
            ("No", "Yes"),
            horizontal=True,
        )

        if veeva_checkbox_value == "Yes":
            ss.veeva_checkbox = True

        st.markdown(
            """
            <style>
            .custom-list {
                color: #7a00e6; /* Change this to your desired color */
                font-size: 16px; /* Adjust the font size if needed */
            }
            .custom-list ul {
                display: flex;
                list-style-type: disc;
                padding-left: 0;
            }
            .custom-list ul li {
                margin-right: 20px; /* Adjust the spacing between items */
            }
            </style>
            <div class="custom-list">
                <h5>For OCCP calculation, we follow Customer Facing Guidance for effort allocation as:</h5>
                <ul>
                    <li>Segment A: 40-50%</li>
                    <li>Segment B: 20-30%</li>
                    <li>Segment C: 10-15%</li>
                </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if not required_fields_filled:
        # or (
        #     ss.is_non_prescriber == "Yes" and not ss.non_prescribers_consent
        # ):
            disabled_review_button = True
            columns = st.columns((5, 5, 5))
            with columns[1]:
                st.warning("**Note:** Please fill all the fields to enable this button")
        else:
            disabled_review_button = False

        columns = st.columns((7, 4, 12, 2))
        with columns[2]:
            review = st.button("Review and Submit", disabled=disabled_review_button)

        if review:
            with st.spinner("Loading your data"):
                obj.create_excel(
                    df, envelope_matrix_df, final_edited_non_prescribers_constraints_df
                )
                data = obj._build_excel_data()
                df_business_constraints, df_brand_specific, df_hcp_constraints = create_business_constraints_file(
                    data=data,
                    sales_line= ss.sales_team,
                    envelope_matrix_df=envelope_matrix_df,
                    df_channel=df_channel,
                    df_sales_line=df_team,
                    df_master=df_master,
                    df_brand=df_brand,
                    df_time=df_time_dim,
                    conn=obj.output_snowflake_con,
                )
                output_table_dict = {
                    "DS_BUSINESS_CONSTRAINTS" : df_business_constraints,
                    "DS_BRAND_SPECIFIC_BUSINESS_CONSTRAINTS" : df_brand_specific,
                    "DS_OCCP_HCP_CONSTRAINTS" : df_hcp_constraints
                }
                obj.review_button(
                    final_hcp_bounds,
                    final_edited_non_prescribers_constraints_df,
                    output_table_dict,
                )


if __name__ == "__main__":
    main()


