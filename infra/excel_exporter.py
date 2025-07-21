"""Excel export service for generating OCCP constraint workbooks."""

import os
from io import BytesIO
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from core.base import ExcelExporter
from core.dto import DTOBundle
from core.errors import ExternalServiceError

class ExcelExporterService(ExcelExporter):
    """Service for generating Excel workbooks from OCCP constraint data."""
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.wb = None
        self.excel_bytes = None
    
    def build(self, bundle: DTOBundle) -> bytes:
        """
        Build Excel workbook from DTO bundle.
        
        Args:
            bundle: Complete DTO bundle with all OCCP constraints
            
        Returns:
            Excel workbook as bytes
        """
        try:
            # Create workbook and worksheet
            self.wb = Workbook()
            ws = self.wb.active
            ws.title = "Business Constraints"
            
            # Add and format headers
            headers = ["Type", "Input", bundle.market.sales_line]
            ws.append(headers)
            self._format_excel_headers(ws)
            
            # Add main business constraints data
            data = self._build_excel_data(bundle)
            for row in data:
                ws.append(row)
            
            self._merge_excel_cells(ws)
            self._format_excel_cells(ws)
            
            # Add envelope matrix sections
            self._add_envelope_matrix_sections(ws, bundle)
            
            # Auto-adjust column widths
            self._auto_adjust_column_width(ws)
            
            # Save workbook to bytes
            output = BytesIO()
            self.wb.save(output)
            output.seek(0)
            self.excel_bytes = output.getvalue()
            
            return self.excel_bytes
            
        except Exception as e:
            raise ExternalServiceError(f"Failed to build Excel workbook: {e}")
    
    def _build_excel_data(self, bundle: DTOBundle) -> list:
        """Build data rows for Excel from DTO bundle."""
        data = [
            ["Cycle", "Generate OCCP for cycle", bundle.cycle.name],
            [
                "OCCP Context",
                "OCCP Cycle Length",
                f"{bundle.cycle.start.strftime('%b %Y')} - {bundle.cycle.end.strftime('%b %Y')}",
            ],
            [
                "OCCP Context",
                "Number of Working Days for Upcoming cycle",
                str(bundle.cycle.working_days),
            ],
            [
                "OCCP Context",
                "Reference Cycle",
                f"{bundle.reference.start.strftime('%b %Y')} - {bundle.reference.end.strftime('%b %Y')}",
            ],
            [
                "OCCP Context",
                "Number of Working Days for Reference cycle",
                str(bundle.reference.working_days),
            ],
            [
                "OCCP Context",
                "Monobrand/Multibrand OCCP",
                "Multibrand" if bundle.market.mode == "Multibrand" else "Monobrand",
            ],
            ["OCCP Context", "OCCP Brand(s)", ", ".join(bundle.market.brands)],
            ["OCCP Context", "OCCP Channels", ", ".join(bundle.capacity.channels)],
        ]
        
        # Add multibrand-specific data
        if bundle.market.mode == "Multibrand":
            data.append(
                [
                    "For Multibrand OCCP only:",
                    "Select which channels can be multibrand interactions?",
                    ", ".join(bundle.capacity.multibrand_channels),
                ]
            )
            
            if bundle.market.specialties:
                specialties_str = ", ".join(
                    [
                        f"{brand_str} : {specialty}"
                        for brand_str, specialty in bundle.market.specialties.items()
                    ]
                )
                data.append(
                    [
                        "For Multibrand OCCP only:",
                        "Specify Specialities that can be promoted together ?",
                        specialties_str,
                    ]
                )
            
            if bundle.distribution:
                brand_distribution_str = ", ".join(
                    [f"{ratio}% {brand}" for brand, ratio in bundle.distribution.ratios.items()]
                )
                data.append(
                    [
                        "For Multibrand OCCP only:",
                        "Brand distribution",
                        brand_distribution_str,
                    ]
                )
        
        # Add capacity constraints
        total_capacity = sum(bundle.capacity.daily_capacity.values())
        data.append(
            [
                "Sales Rep. Constraints",
                "Avg Rep Capacity for All Channels",
                f"{total_capacity:.2f}",
            ]
        )
        
        for channel, capacity in bundle.capacity.daily_capacity.items():
            if bundle.capacity.non_prescriber_included:
                label = f"Avg Rep Capacity per day for {channel} (Prescriber and Non-Prescriber Combined)"
            else:
                label = f"Avg Rep Capacity per day for {channel}"
            
            data.append(
                [
                    "Sales Rep. Constraints",
                    label,
                    f"{capacity:.2f}",
                ]
            )
        
        # Add e-consent flag if RTE channels are present
        rte_channels = ["RTE-Open", "RTE-Sent"]
        if any(ch in bundle.capacity.channels for ch in rte_channels):
            consent_flag = "Yes" if bundle.capacity.e_consent_rte else "No"
            data.append(
                [
                    "Additional Constraints",
                    "eConsent required for RTE?",
                    consent_flag,
                ]
            )
        
        return data
    
    def _add_envelope_matrix_sections(self, ws, bundle: DTOBundle):
        """Add envelope matrix sections to Excel."""
        # Add historical envelope matrix
        if bundle.envelopes_hist:
            self._add_historical_envelope_matrix(ws, bundle.envelopes_hist)
        
        # Add segment envelope matrix
        if bundle.envelopes_seg:
            self._add_segment_envelope_matrix(ws, bundle.envelopes_seg)
        
        # Add non-prescriber envelope matrix
        if bundle.non_prescriber:
            self._add_non_prescriber_envelope_matrix(ws, bundle.non_prescriber)
    
    def _add_historical_envelope_matrix(self, ws, envelopes):
        """Add historical envelope matrix to Excel."""
        # Group by channel
        from collections import defaultdict
        channel_groups = defaultdict(list)
        for envelope in envelopes:
            channel_groups[envelope.channel].append(envelope)
        
        for channel, channel_envelopes in channel_groups.items():
            # Blank row for spacing
            ws.append([""])
            
            # Section header
            ws.append([f"Channel: {channel}"])
            header_row_idx = ws.max_row
            self._format_section_header(ws, header_row_idx, 4)
            
            # Matrix headers
            matrix_headers = ["Type", "REFERENCE_CYCLE_ACTUAL", "Min_Value", "Max_Value"]
            ws.append(matrix_headers)
            self._format_matrix_headers(ws, ws.max_row, 4)
            
            # Data rows
            start_row = ws.max_row + 1
            for envelope in channel_envelopes:
                ws.append([
                    "", 
                    envelope.reference_cycle_actual,
                    envelope.rule.min_val,
                    envelope.rule.max_val
                ])
            end_row = ws.max_row
            
            # Merge type column and add table
            self._merge_type_column(ws, start_row, end_row, "HCP Constraints")
            self._add_excel_table(ws, start_row, end_row, f"Table_{channel}_{start_row}")
    
    def _add_segment_envelope_matrix(self, ws, envelopes):
        """Add segment envelope matrix to Excel."""
        # Group by channel and brand
        from collections import defaultdict
        group_dict = defaultdict(list)
        for envelope in envelopes:
            key = (envelope.channel, envelope.brand)
            group_dict[key].append(envelope)
        
        for (channel, brand), brand_envelopes in group_dict.items():
            # Blank row for spacing
            ws.append([""])
            
            # Section header
            ws.append([f"Channel: {channel} | Brand: {brand}"])
            header_row_idx = ws.max_row
            self._format_section_header(ws, header_row_idx, 5)
            
            # Matrix headers
            matrix_headers = ["Type", "BRAND", "SEGMENT", "Min_Value", "Max_Value"]
            ws.append(matrix_headers)
            self._format_matrix_headers(ws, ws.max_row, 5)
            
            # Data rows
            start_row = ws.max_row + 1
            for envelope in brand_envelopes:
                ws.append([
                    "",
                    envelope.brand,
                    envelope.segment,
                    envelope.rule.min_val,
                    envelope.rule.max_val
                ])
            end_row = ws.max_row
            
            # Merge type column and add table
            self._merge_type_column(ws, start_row, end_row, "HCP Constraints")
            self._add_excel_table(ws, start_row, end_row, f"Table_{channel}_{brand}_{start_row}")
    
    def _add_non_prescriber_envelope_matrix(self, ws, envelopes):
        """Add non-prescriber envelope matrix to Excel."""
        # Blank row for spacing
        ws.append([""])
        
        # Section header
        ws.append(["NON-PRESCRIBERS ENVELOPE RULES"])
        header_row_idx = ws.max_row
        self._format_section_header(ws, header_row_idx, 4)
        
        # Matrix headers
        matrix_headers = ["Type", "CHANNEL", "Min_Value", "Max_Value"]
        ws.append(matrix_headers)
        self._format_matrix_headers(ws, ws.max_row, 4)
        
        # Data rows
        start_row = ws.max_row + 1
        for envelope in envelopes:
            ws.append([
                "",
                envelope.channel,
                envelope.rule.min_val,
                envelope.rule.max_val
            ])
        end_row = ws.max_row
        
        # Merge type column and add table
        self._merge_type_column(ws, start_row, end_row, "NON-PRESCRIBERS Constraints")
        self._add_excel_table(ws, start_row, end_row, f"Table_Non_Prescribers_{start_row}")
    
    def _format_excel_headers(self, ws):
        """Format Excel headers."""
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="7a00e6", end_color="7a00e6", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
    
    def _format_section_header(self, ws, row_idx, num_cols):
        """Format section header row."""
        for cell in ws[row_idx]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color="dcd0ff", end_color="dcd0ff", fill_type="solid")
            cell.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
        ws.merge_cells(
            start_row=row_idx,
            start_column=1,
            end_row=row_idx,
            end_column=num_cols,
        )
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    def _format_matrix_headers(self, ws, row_idx, num_cols):
        """Format matrix header row."""
        for cell in ws[row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7a00e6", end_color="dcd0ff", fill_type="solid")
            cell.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
    
    def _merge_type_column(self, ws, start_row, end_row, type_value):
        """Merge the type column and set its value."""
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=end_row,
            end_column=1,
        )
        ws.cell(row=start_row, column=1).value = type_value
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    def _add_excel_table(self, ws, start_row, end_row, table_name):
        """Add Excel table style."""
        tab = Table(
            displayName=table_name,
            ref=f"A{start_row}:D{end_row}",
        )
        style = TableStyleInfo(
            name="TableStyleLight12",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
    
    def _merge_excel_cells(self, ws):
        """Merge cells in Excel."""
        col = 1
        start_row = 2
        
        while start_row <= ws.max_row:
            current_value = ws.cell(start_row, col).value
            
            if not current_value:
                start_row += 1
                continue
            
            end_row = start_row + 1
            while (
                end_row <= ws.max_row and ws.cell(end_row, col).value == current_value
            ):
                end_row += 1
            
            if end_row - start_row > 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row - 1,
                    end_column=col,
                )
                ws.cell(start_row, col).alignment = Alignment(vertical="top")
            
            start_row = end_row
    
    def _format_excel_cells(self, ws):
        """Format Excel cells."""
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )
    
    def _auto_adjust_column_width(self, ws):
        """Auto-adjust column widths."""
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2 